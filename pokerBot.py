from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import tkinter as tk
import os
from openpyxl import Workbook, load_workbook
import time
import pyautogui
from tkinter import messagebox
import threading




# Caminhos para o geckodriver e perfil do Firefox
geckodriver_path = os.path.join(os.path.dirname(__file__), 'geckodriver.exe')
profile_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'profile.txt')


# Lê o caminho do perfil a partir do arquivo de texto
try:
    with open(profile_file_path, 'r') as file:
        profile_path = file.readline().strip()  # Lê a primeira linha e remove espaços em branco
    print(f"Caminho do perfil do Firefox: {profile_path}")
except FileNotFoundError:
    print("Arquivo 'profile.txt' não encontrado.")
except Exception as e:
    print(f"Ocorreu um erro: {e}")



# Configurar o serviço e perfil do Firefox
service = Service(geckodriver_path)
options = Options()
options.profile = profile_path

driver = webdriver.Firefox(service=service, options=options)

url = 'https://app.gtowizard.com/'
driver.get(url)









def perform_scroll(scroll_value):
    time.sleep(0.5)
    pyautogui.keyDown('ctrl')
    pyautogui.scroll(scroll_value)
    time.sleep(0.5)
    pyautogui.keyUp('ctrl')




adjustment_complete = threading.Event()
# Função para contar as divs e ajustar o zoom
def count_and_adjust_divs():
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.gw_table_body_row.repfloptblrow.gw_table_body_row_hoverable.gw_hvr.gw_hvr_pcn.cursor-pointer'))
        )
        
        containers = driver.find_elements(By.CSS_SELECTOR, '.gw_table_body_row.repfloptblrow.gw_table_body_row_hoverable.gw_hvr.gw_hvr_pcn.cursor-pointer')
        num_containers = len(containers)

        print(num_containers)
        while num_containers < 105 or num_containers >= 120:
            if num_containers < 105:
                perform_scroll(-80)  # Zoom in
            elif num_containers >= 120:
                perform_scroll(80)  # Zoom out
            
            time.sleep(1)  # Aguarde um momento para que o efeito do scroll se aplique
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.gw_table_body_row.repfloptblrow.gw_table_body_row_hoverable.gw_hvr.gw_hvr_pcn.cursor-pointer'))
            )
            containers = driver.find_elements(By.CSS_SELECTOR, '.gw_table_body_row.repfloptblrow.gw_table_body_row_hoverable.gw_hvr.gw_hvr_pcn.cursor-pointer')
            num_containers = len(containers)

        adjustment_complete.set()
        
    except Exception as e:
       print(f"Erro inesperado. {str(e)}")



def start_count_and_adjust():
    thread = threading.Thread(target=count_and_adjust_divs)
    thread.start()



def extract_header_info():
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".gw_table_head_row"))
        )

        cabecalho = driver.find_element(By.CSS_SELECTOR, ".gw_table_head_row")

        cabecalho_texto = []
        celulas = cabecalho.find_elements(By.CLASS_NAME, "gw_table_head_cell_content")
        for celula in celulas:
            # cabecalho_texto.append(celula.text.strip())
            texto = celula.text.strip()
            if texto != "Estratégia":
                cabecalho_texto.append(texto)

        
        print(f"Cabeçalho: {', '.join(cabecalho_texto)}")
        return [cabecalho_texto]  # Retornar como lista de lista para manter a consistência

    except Exception as e:
        print(f"Erro ao extrair informações do cabeçalho: {e}")
        messagebox.showerror("Erro", f"Erro ao extrair informações do cabeçalho: {e}")
        return []

def extract_lines_info():
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".f-column.large-scroll-track .gw_table_body_row"))
        )

        containers = driver.find_elements(By.CSS_SELECTOR, ".f-column.large-scroll-track .gw_table_body_row")

        linhas_texto = []
        for i, container in enumerate(containers, start=1):
            linha = container.text.strip().split('\n')  # Separar as linhas em listas de strings
            linhas_texto.append(linha)
            print(f"Linha {i}: {linha}")

        return linhas_texto

    except Exception as e:
        print(f"Erro ao extrair informações: {e}")
        return []





# Função para extrair informações das três primeiras cartas
def extract_first_three_cards_info(container):
    card_info = []
    try:
        card_blocks = container.find_elements(By.CSS_SELECTOR, ".cardsymbols_block")[:3]
        for card_block in card_blocks:
            card_value = card_block.find_element(By.CSS_SELECTOR, ".cardsymbols_value").text.strip()
            card_symbol_element = card_block.find_element(By.CSS_SELECTOR, ".cardsymbols_symbol svg")
            svg_path = card_symbol_element.find_element(By.CSS_SELECTOR, 'path').get_attribute('d')
            if "M3.89" in svg_path:
                suit_name = "PAUS"
            elif "M30.97" in svg_path:
                suit_name = "COPAS"
            elif "M28.29" in svg_path:
                suit_name = "OUROS"
            else:
                suit_name = "ESPADAS"
            card_info.append(f"{card_value} de {suit_name}")
        return card_info
    except Exception as e:
        print(f"Erro ao extrair informações das três primeiras cartas: {e}")
        return None




# Função para salvar no excel
def save_to_excel(workbook, sheet_name, all_data_to_print):
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook[sheet_name]
    first_empty_row = sheet.max_row + 1

    for row_data in all_data_to_print:
        sheet.append(row_data)
    workbook.save(excel_file_path)





# Função para extrair outras informações únicas ignorando textos e capturando apenas números
def extract_other_info(div_pai):
    try:
        # Encontrar todas as divs filhas com as classes específicas dentro da div pai encontrada
        divs_filhas_position = div_pai.find_elements(By.XPATH, ".//div[@class='position-absolute w-100 f-center']")
        divs_filhas_text_center = div_pai.find_elements(By.XPATH, ".//div[@class='text-center']")

        other_info = []
        for div_filha in divs_filhas_position + divs_filhas_text_center:
            # Obter o texto da div filha
            texto = div_filha.text.strip()
            other_info.append(texto)

        return other_info

    except Exception as e:
        print(f"Erro ao extrair informações: {str(e)}")
        return []
    



def show_dialog():
    root.lift()
    root.focus_force()     


def bring_dialog_to_front():
    root.after(2 * 1000, show_dialog) 


# Função principal para extrair e salvar todas as informações
def extract_and_save_all_info(driver, excel_file_path, sheet_name):
    all_data_to_print = []
    previous_first_data = None
    # Ler dados existentes
    existing_data = set()
    if os.path.exists(excel_file_path):
        workbook = load_workbook(excel_file_path)
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            existing_data.add(tuple(row))  # Adiciona como tupla para facilitar a verificação

    try:
        while True:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.gw_table_body_row.repfloptblrow.gw_table_body_row_hoverable.gw_hvr.gw_hvr_pcn.cursor-pointer'))
            )
            containers = driver.find_elements(By.CSS_SELECTOR, '.gw_table_body_row.repfloptblrow.gw_table_body_row_hoverable.gw_hvr.gw_hvr_pcn.cursor-pointer')

            print(f"Número de containers encontrados: {len(containers)}")

            # Verifique se o arquivo existe, caso contrário, crie um novo
            if not os.path.exists(excel_file_path):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = sheet_name
                workbook.save(excel_file_path)
            else:
                workbook = load_workbook(excel_file_path)

            # Contador para controlar as iterações
            iteration_count = 0 
            data_changed = False

            for i, container in enumerate(containers, start=1):
                iteration_count += 1

                xpath = f"(//div[@class='gw_table_body_row repfloptblrow gw_table_body_row_hoverable gw_hvr gw_hvr_pcn cursor-pointer'])[{i}]"
                div_pai = driver.find_element(By.XPATH, xpath)

                first_three_cards_info = extract_first_three_cards_info(container)
                if first_three_cards_info:
                    cards_info = ' - '.join(first_three_cards_info)

                other_info = extract_other_info(div_pai)

                current_data = [cards_info] + other_info

                if current_data == previous_first_data:
                    print("Dados repetidos. Encerrando extração.")
                    print("CURRENT:", current_data) 
                    print("PREVIOUS:", previous_first_data) 
                    data_changed = True
                    break

                # Verificar se current_data já está na planilha
                if tuple(current_data) not in existing_data:
                    all_data_to_print.append(current_data)
                    existing_data.add(tuple(current_data))  # Adiciona à lista de dados existentes

                if iteration_count == 1:
                    previous_first_data = current_data
                    print("FIRST:", previous_first_data)

            if data_changed:
                break

            save_to_excel(workbook, sheet_name, all_data_to_print)
            print("SALVAMENTO CHAMADO!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")

            all_data_to_print = []

            pyautogui.click() 
            # Scroll down
            if iteration_count % len(containers) == 0:
                for _ in range(24):
                    pyautogui.scroll(-100)
                time.sleep(0.5)



        bring_dialog_to_front()

        return True, "Informações extraídas e salvas com sucesso."

    except Exception as e:
        print(f"Erro ao extrair informações: {e}")
        return False, f"Erro ao extrair informações: {e}"




def btn_call():
    time.sleep(5)
    # print("INICIANDO SEQUENCIA!")
    global adjustment_complete
    adjustment_complete.clear()
    threading.Thread(target=count_and_adjust_divs).start()
    adjustment_complete.wait()
    global excel_file_path
    label_error.config(text="")
    file_name = entry.get()
    if not file_name:
        label_error.config(text="Insira um nome para o arquivo Excel:", fg="red")
        return
    base_dir = os.path.abspath('planilhas')
    excel_file_path = os.path.join(base_dir, f"{file_name}.xlsx")
    sheet_name = entry_sheet_name.get()
    if not sheet_name:
        label_error.config(text="Insira um nome para a nova aba:", fg="red")
        return
    

    cabecalho_texto = extract_header_info()
    linhas_texto = extract_lines_info()
    if not cabecalho_texto or not linhas_texto:
        label_error.config(text="Erro ao extrair informações.", fg="red")
        return
    
    all_data_to_print = cabecalho_texto + linhas_texto
    # Abra o workbook ou crie um novo se não existir
    try:
        workbook = load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.active.title = sheet_name

    save_to_excel(workbook, sheet_name, all_data_to_print)

    success, message = extract_and_save_all_info(driver, excel_file_path, sheet_name)
    if success:
        label_error.config(text=message, fg="green")
    else:
        label_error.config(text=message, fg="red")




# Caixa de diálogo
def config_handler():
    threading.Thread(target=btn_call).start()




# Configurar a interface gráfica com tkinter
root = tk.Tk()
root.title("Configurações")
root.geometry("400x250")

label_file_name = tk.Label(root, text="Nome do arquivo Excel:")
label_file_name.pack(pady=(20, 5))

entry = tk.Entry(root, width=30)
entry.pack(pady=5)

label_sheet_name = tk.Label(root, text="Nome da nova aba:")
label_sheet_name.pack(pady=(20, 5))

entry_sheet_name = tk.Entry(root, width=30)
entry_sheet_name.pack(pady=5)

label_error = tk.Label(root, text="", fg="red")
label_error.pack(pady=10)

button = tk.Button(root, text="Tudo Pronto!", command=config_handler)
button.pack(pady=5)

root.mainloop()
